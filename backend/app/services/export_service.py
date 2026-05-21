import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExportService:
    @staticmethod
    def generate_excel_report(logs: list[dict], title: str = "Attendance Report") -> io.BytesIO:
        """
        Convert a list of attendance log rows into a styled Excel file.
        Returns a BytesIO binary stream of the generated workbook.
        """
        # 1. Transform raw dict rows into clean DataFrame
        df_data = []
        for log in logs:
            df_data.append({
                "Employee ID": log.get("employee_id"),
                "Employee Name": log.get("employee_name"),
                "Department": log.get("department_name", "N/A"),
                "Date": log.get("date"),
                "Check-in Time": log.get("check_in_time") or "-",
                "Check-out Time": log.get("check_out_time") or "-",
                "Attendance Status": log.get("attendance_status"),
                "Late Minutes": log.get("late_minutes", 0),
                "Total Work Hours": log.get("work_duration", 0.0)
            })

        df = pd.DataFrame(df_data)

        # 2. Write to openpyxl via pandas ExcelWriter
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Attendance Data", index=False)
            
            # Access workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets["Attendance Data"]
            
            # 3. Add Premium corporate styling using openpyxl
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Charcoal
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
            
            # Style header row
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            # Style data rows and formats
            font_regular = Font(name="Segoe UI", size=10)
            fill_present = PatternFill(start_color="DEF7EC", end_color="DEF7EC", fill_type="solid") # Light green
            fill_late = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light amber
            fill_absent = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid") # Light red
            
            for row_idx in range(2, worksheet.max_row + 1):
                # Fetch attendance status to highlight cells
                status_cell = worksheet.cell(row=row_idx, column=7) # Column 7 is 'Attendance Status'
                status = status_cell.value
                
                row_fill = None
                if status == "Present":
                    row_fill = fill_present
                elif status in ["Late", "Early Checkout", "Half Day"]:
                    row_fill = fill_late
                elif status == "Absent":
                    row_fill = fill_absent
                
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = font_regular
                    cell.border = thin_border
                    
                    # Alignments
                    if col_idx in [4, 5, 6, 7]: # Date, Check-in, Check-out, Status
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                    
                    # Numbers Formatting
                    if col_idx == 8: # Late Minutes
                        cell.number_format = '#,##0'
                    elif col_idx == 9: # Work Hours
                        cell.number_format = '0.00'
                        
                    # Apply row highlights based on status
                    if row_fill and col_idx == 7: # Only highlight the status column for clean readability
                        cell.fill = row_fill

            # 4. Auto-fit column widths
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output.seek(0)
        return output
